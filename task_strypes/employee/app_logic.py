from openpyxl import load_workbook
from io import BytesIO

from datetime import datetime
import re

MOBILE_REGEX = re.compile(r'\d+')
POSITION_FIELD_REPRESENTATION = {'ceo': 'CEO', 'junior developer': 'junior_dev',
                                 'senior developer': 'senior_dev', 'team lead': 'team_lead',
                                 'project manager': 'project_manager'}


def read_xlsx_from_memory(xlsx_file):
    file_in_memory = xlsx_file.FILES['file_xlsx'].read()
    wb = load_workbook(filename=BytesIO(file_in_memory))
    ws = wb.active
    return ws


def get_data_from_xlsx_sheet(ws):
    employees = []
    row = 2
    columns = ['first_name', 'last_name', 'mobile_num', 'start_date', 'position', 'salary', 'employee_id']
    while ws.cell(row=row, column=1).value:
        employee = []
        for i in range(1, 8):
            employee.append(ws.cell(row=row, column=i).value)

        employees.append(
            dict(zip(columns, employee))
        )
        row += 1

    return employees


def format_start_date_field(employee):
    if type(employee.get('start_date')) == str:
        formatted_date = datetime.strptime(employee.get('start_date'), '%d/%m/%Y').date()
    else:
        formatted_date = employee.get('start_date').date()

    return formatted_date


def format_ws_fields(employees):
    for employee in employees:
        formatted_date_filed = format_start_date_field(employee)
        formatted_position_field = POSITION_FIELD_REPRESENTATION.get(employee.get('position').lower())
        salary_extracted = int(re.search(MOBILE_REGEX, employee.get('salary')).group(0))

        employee.update({'start_date': formatted_date_filed,
                         'salary': salary_extracted,
                         'position': formatted_position_field}
                        )


def create_update_data(request, employee_to_update):
    first_name = request.data.get('first_name') \
        if request.data.get('first_name') else employee_to_update.first_name
    last_name = request.data.get('last_name') \
        if request.data.get('last_name') else employee_to_update.last_name
    mobile_num = int(request.data.get('mobile_num')) \
        if request.data.get('mobile_num') else employee_to_update.mobile_num
    start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d').date() \
        if request.data.get('start_date') else employee_to_update.start_date
    position = POSITION_FIELD_REPRESENTATION.get(request.data.get('position').lower()) \
        if request.data.get('position') else employee_to_update.position
    salary = int(request.data.get('salary')) \
        if request.data.get('salary') else employee_to_update.salary
    employee_id = request.data.get('employee_id') \
        if request.data.get('employee_id') else employee_to_update.employee_id

    return {'first_name': first_name,
            'last_name': last_name,
            'mobile_num': mobile_num,
            'start_date': start_date,
            'position': position,
            'salary': salary,
            'employee_id': employee_id,
            }
