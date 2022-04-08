from rest_framework.views import APIView
from rest_framework import serializers, status
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.response import Response
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

from django.views.generic.edit import CreateView
from django.http import HttpResponse
import os

from employee.models import Employee
import re

from django.http import HttpResponseBadRequest, HttpResponse
from django.urls import reverse

MOBILE_REGEX = re.compile(r'\d+')


class Home(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'home_page.html'

    def get(self, request):
        return Response(status=status.HTTP_200_OK)


class EmployeesApi(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'all_employees.html'

    def get(self, request):
        queryset = Employee.objects.all()
        return Response({'profiles': queryset})


class UploadXLSX(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'file_upload.html'

    class EmployeeSerializer(serializers.Serializer):
        first_name = serializers.CharField()
        last_name = serializers.CharField()
        mobile_num = serializers.IntegerField()
        start_date = serializers.DateField()
        position = serializers.CharField()
        salary = serializers.IntegerField()
        employee_id = serializers.CharField()

    def get(self, request):
        return Response()

    def post(self, request):

        if request.FILES['file_xlsx']:
            file_in_memory = request.FILES['file_xlsx'].read()
            wb = load_workbook(filename=BytesIO(file_in_memory))
            # wb.iso_dates = True

            ws = wb.active
            var = ws.cell(row=1, column=2)

            employees = []
            n = 2
            columns = ['first_name', 'last_name', 'mobile_num', 'start_date', 'position', 'salary', 'employee_id']
            while ws.cell(row=n, column=1).value:
                employee = []

                for i in range(1, 8):
                    employee.append(ws.cell(row=n, column=i).value)

                employees.append(
                    dict(zip(columns, employee))
                )
                n += 1

            POSITION_FIELD_REPRESENTATION = {'ceo': 'CEO', 'junior developer': 'junior_dev',
                                             'senior developer': 'senior_dev', 'team lead': 'team_lead',
                                             'project manager': 'project_manager'}
            for i in employees:
                if type(i.get('start_date')) == str:
                    formatted_date = datetime.strptime(i.get('start_date'), '%d/%m/%Y').date()
                else:
                    formatted_date = i.get('start_date').date()

                position_db_represented = POSITION_FIELD_REPRESENTATION.get(i.get('position').lower())

                salary_extracted = int(re.search(MOBILE_REGEX, i.get('salary')).group(0))

                i.update({'start_date': formatted_date,
                          'salary': salary_extracted,
                          'position': position_db_represented}
                         )






            serializer_employees = self.EmployeeSerializer(data=employees, many=True)

            if serializer_employees.is_valid():
                table_entries = serializer_employees.data

                db_bulk_data = [Employee(first_name=table_entry.get('first_name'),
                                         last_name=table_entry.get('last_name'),
                                         mobile_num=table_entry.get('mobile_num'),
                                         start_date=table_entry.get('start_date'),
                                         position=table_entry.get('position'),
                                         salary=table_entry.get('salary'),
                                         employee_id=table_entry.get('employee_id'),
                                         )
                                for table_entry in table_entries]

                Employee.objects.bulk_create(db_bulk_data)
        import pdb; pdb.set_trace()


class AddEmployee(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'create_employee.html'

    class EmployeeSerializer(serializers.Serializer):       # DRY principle
        first_name = serializers.CharField()
        last_name = serializers.CharField()
        mobile_num = serializers.IntegerField()
        start_date = serializers.DateField()
        position = serializers.CharField()
        salary = serializers.IntegerField()
        employee_id = serializers.CharField()

    def get(self, request):
        return Response()

    def post(self, request):
        # DRY
        POSITION_FIELD_REPRESENTATION = {'ceo': 'CEO', 'junior developer': 'junior_dev',
                                         'senior developer': 'senior_dev', 'team lead': 'team_lead',
                                         'project manager': 'project_manager'}

        employee = {'first_name': request.data.get('first_name'),
                    'last_name': request.data.get('last_name'),
                    'mobile_num': int(request.data.get('mobile_num')),
                    'start_date': datetime.strptime(request.data.get('start_date'), '%Y-%m-%d').date(),
                    'position': POSITION_FIELD_REPRESENTATION.get(request.data.get('position').lower()),
                    'salary': int(request.data.get('salary')),
                    'employee_id': request.data.get('employee_id'),
                    }

        serializer_employees = self.EmployeeSerializer(data=employee)

        if serializer_employees.is_valid():
            Employee.objects.create(first_name=employee.get('first_name'),
                                    last_name=employee.get('last_name'),
                                    mobile_num=employee.get('mobile_num'),
                                    start_date=employee.get('start_date'),
                                    position=employee.get('position'),
                                    salary=employee.get('salary'),
                                    employee_id=employee.get('employee_id'),
                                    )


class EditEmployee(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'update_employee.html'

    class EmployeeSerializer(serializers.Serializer):       # DRY principle
        first_name = serializers.CharField()
        last_name = serializers.CharField()
        mobile_num = serializers.IntegerField()
        start_date = serializers.DateField()
        position = serializers.CharField()
        salary = serializers.IntegerField()
        employee_id = serializers.CharField()

    def get(self, request):
        return Response()

    def post(self, request):

        employee_to_update_set = Employee.objects.filter(employee_id=request.data.get('id_to_update'))

        if employee_to_update_set:
            employee_to_update = employee_to_update_set[0]
        # DRY
            POSITION_FIELD_REPRESENTATION = {'ceo': 'CEO', 'junior developer': 'junior_dev',
                                             'senior developer': 'senior_dev', 'team lead': 'team_lead',
                                             'project manager': 'project_manager'}

            first_name = request.data.get('first_name') \
                if request.data.get('first_name') else employee_to_update.first_name
            last_name = request.data.get('last_name') \
                if request.data.get('last_name') else employee_to_update.last_name
            mobile_num = int(request.data.get('mobile_num')) \
                if request.data.get('mobile_num') else employee_to_update.mobile_num
            start_date = datetime.strptime(request.data.get('start_date'), '%Y-%m-%d').date() \
                if request.data.get('start_date') else employee_to_update.start_date
            position = POSITION_FIELD_REPRESENTATION.get(request.data.get('position').lower()) \
                if request.data.get('position') else employee_to_update.position
            salary = int(request.data.get('salary')) \
                if request.data.get('salary') else employee_to_update.salary
            employee_id = request.data.get('employee_id') \
                if request.data.get('employee_id') else employee_to_update.employee_id

            employee = {'first_name': first_name,
                        'last_name': last_name,
                        'mobile_num': mobile_num,
                        'start_date': start_date,
                        'position': position,
                        'salary': salary,
                        'employee_id': employee_id,
                        }

            serializer_employees = self.EmployeeSerializer(data=employee)

            if serializer_employees.is_valid():
                employee_to_update_set.update(first_name=employee.get('first_name'),
                                              last_name=employee.get('last_name'),
                                              mobile_num=employee.get('mobile_num'),
                                              start_date=employee.get('start_date'),
                                              position=employee.get('position'),
                                              salary=employee.get('salary'),
                                              employee_id=employee.get('employee_id')

                                              )
                message = f'Employee with ID {request.data.get("id_to_update")} has been updated \n ' \
                          f'<p><a href="{reverse("employees")}">Home Page</a></p>'
                return HttpResponse(message, status=status.HTTP_201_CREATED)

        message = f'Employee with ID {request.data.get("id_to_update")} NOT found'
        return HttpResponseBadRequest(message, status=status.HTTP_404_NOT_FOUND)

